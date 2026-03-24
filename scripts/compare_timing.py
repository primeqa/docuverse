#!/usr/bin/env python3
"""Compare timing statistics across multiple model runs.

Takes a list of .timing.json files and timing keys, produces an Excel
spreadsheet with one tab per timing key, listing timing statistics for
each model. Statistics can be divided by a batch size (auto-detected
from .metrics files if not provided).

Usage:
    python scripts/compare_timing.py \
        --timing_files output/model1.timing.json output/model2.timing.json \
        --keys "ingest_and_test::ingest::encode" "ingest_and_test::search::retrieve::encode::model_forward" \
        --output comparison.xlsx \
        --batch_size 128

    # Auto-detect batch size from .metrics files:
    python scripts/compare_timing.py \
        --timing_files output/model1.timing.json output/model2.timing.json \
        --keys "ingest_and_test::ingest::encode" \
        --output comparison.xlsx
"""

import argparse
import json
import os
import re
import sys
import yaml

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def parse_args():
    parser = argparse.ArgumentParser(
        description="Compare timing statistics across model runs and produce an Excel spreadsheet."
    )
    parser.add_argument(
        "--timing_files", "-t", nargs="+", required=True,
        help="List of .timing.json files to compare"
    )
    parser.add_argument(
        "--keys", "-k", nargs="+", required=True,
        help="Timing keys to extract (e.g. 'ingest_and_test::ingest::encode')"
    )
    parser.add_argument(
        "--output", "-o", default="timing_comparison.xlsx",
        help="Output Excel file (default: timing_comparison.xlsx)"
    )
    parser.add_argument(
        "--batch_size", "-b", type=int, default=None,
        help="Batch size to divide statistics by. If not provided, auto-detect from .metrics files."
    )
    parser.add_argument(
        "--per_item", action="store_true", default=False,
        help="Divide statistics by batch size to show per-item timings"
    )
    parser.add_argument(
        "--stats", "-s", nargs="+",
        default=["count", "sum", "mean", "median", "std", "p95", "p99", "max"],
        help="Statistics to include in the output"
    )
    return parser.parse_args()


def extract_model_name(timing_file):
    """Extract a short model name from the timing file path."""
    basename = os.path.basename(timing_file)
    # Strip .timing.json suffix
    for suffix in [".timing.json", ".json"]:
        if basename.endswith(suffix):
            basename = basename[:-len(suffix)]
            break
    # Strip common prefixes for readability
    for prefix in ["unified-search-full-", "unified-search-short23k-", "unified-search-"]:
        if basename.startswith(prefix):
            basename = basename[len(prefix):]
            break
    return basename


def read_metrics_config(metrics_file):
    """Read the config dict from a .metrics file's config section.

    The config is written as a YAML string with literal '\\n' for newlines,
    on the line(s) following '****** Config: *******'.
    Returns the parsed config dict, or None if not found/parseable.
    """
    if not os.path.exists(metrics_file):
        return None

    with open(metrics_file, "r") as f:
        lines = f.readlines()

    # Find the config section
    config_start = None
    for i, line in enumerate(lines):
        if "****** Config: *******" in line:
            config_start = i + 1
            break

    if config_start is None:
        return None

    # Collect config lines until the next separator or end of file
    config_text = ""
    for i in range(config_start, len(lines)):
        line = lines[i].strip()
        if line.startswith("=" * 5):
            break
        config_text += line

    # The config has literal \n instead of actual newlines
    config_yaml = config_text.replace("\\n", "\n")

    try:
        config = yaml.safe_load(config_yaml)
    except yaml.YAMLError:
        return None

    return config if isinstance(config, dict) else None


def get_metrics_info(timing_file, default_batch_size=None):
    """Get batch_size and max_doc_length for a timing file from the .metrics config.

    Returns (batch_size, max_doc_length).
    """
    metrics_file = timing_file.replace(".timing.json", ".metrics")
    config = read_metrics_config(metrics_file)

    batch_size = default_batch_size
    max_doc_length = None

    if config is not None:
        retriever = config.get("retriever", {})
        if isinstance(retriever, dict):
            if batch_size is None:
                batch_size = retriever.get("bulk_batch")
            max_doc_length = retriever.get("max_doc_length")

    return batch_size, max_doc_length


def load_timing_data(timing_file):
    """Load timing data from a .timing.json file."""
    with open(timing_file, "r") as f:
        return json.load(f)


def shorten_key_for_tab(key):
    """Shorten a timing key for use as an Excel tab name (max 31 chars)."""
    # Take the last 2-3 components, use '.' separator (Excel forbids ':')
    parts = key.split("::")
    short = ".".join(parts[-2:]) if len(parts) > 2 else key.replace("::", ".")
    # Excel sheet names cannot contain: : \ / ? * [ ]
    for ch in [":", "\\", "/", "?", "*", "[", "]"]:
        short = short.replace(ch, "_")
    if len(short) > 31:
        short = short[-31:]
    return short


_table_counter = 0

def write_sheet(ws, key, timing_files, timing_data_list, batch_sizes, max_doc_lengths, stats):
    """Write one sheet for a timing key, formatted as an Excel Table with color scales."""
    global _table_counter
    _table_counter += 1

    num_format = "#,##0.000"
    int_format = "#,##0"

    # Column definitions: (header_name, width)
    fixed_headers = [("Model", 50), ("Max Tokens", 12), ("Batch Size", 12)]
    all_headers = fixed_headers + [(s, 14) for s in stats]
    stat_start_col = len(fixed_headers) + 1
    total_cols = len(all_headers)

    # Row 1: header
    header_row = 1
    for col_idx, (name, _) in enumerate(all_headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=name)

    # Data rows
    n_files = len(timing_files)
    for i, timing_file in enumerate(timing_files):
        row = header_row + 1 + i
        model_name = extract_model_name(timing_file)
        data = timing_data_list[i]
        batch_size = batch_sizes[i]
        max_doc_length = max_doc_lengths[i]

        ws.cell(row=row, column=1, value=model_name)

        if max_doc_length is not None:
            ws.cell(row=row, column=2, value=max_doc_length).number_format = int_format
        else:
            ws.cell(row=row, column=2, value="-")

        if batch_size is not None:
            ws.cell(row=row, column=3, value=batch_size).number_format = int_format
        else:
            ws.cell(row=row, column=3, value="-")

        statistics = data.get("statistics", {})
        key_stats = statistics.get(key)

        if key_stats is None:
            for col_offset in range(len(stats)):
                ws.cell(row=row, column=stat_start_col + col_offset, value="N/A")
            continue

        divisor = batch_size if batch_size is not None else 1
        for col_offset, stat in enumerate(stats):
            col = stat_start_col + col_offset
            value = key_stats.get(stat)
            if value is not None:
                if stat in ("count", "sum"):
                    fmt = int_format if stat == "count" else num_format
                    ws.cell(row=row, column=col, value=value).number_format = fmt
                elif stat == "std":
                    ws.cell(row=row, column=col, value=value / divisor).number_format = num_format
                else:
                    ws.cell(row=row, column=col, value=value / divisor).number_format = num_format
            else:
                ws.cell(row=row, column=col, value="N/A")

    # Format as Excel Table
    last_row = header_row + max(n_files, 1)
    last_col_letter = get_column_letter(total_cols)
    table_ref = f"A{header_row}:{last_col_letter}{last_row}"
    table_name = f"Timing_{_table_counter}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    # Green-to-red color scale on mean and median columns
    if n_files >= 2:
        for col_offset, stat in enumerate(stats):
            if stat in ("mean", "median"):
                col_letter = get_column_letter(stat_start_col + col_offset)
                data_range = f"{col_letter}{header_row + 1}:{col_letter}{last_row}"
                ws.conditional_formatting.add(
                    data_range,
                    ColorScaleRule(
                        start_type="min", start_color="63BE7B",  # green
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
                        end_type="max", end_color="F8696B",  # red
                    )
                )

    # Column widths
    for col_idx, (_, width) in enumerate(all_headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main():
    args = parse_args()

    # Load all timing data
    timing_data_list = []
    batch_sizes = []
    max_doc_lengths = []
    for tf in args.timing_files:
        if not os.path.exists(tf):
            print(f"Warning: {tf} not found, skipping.", file=sys.stderr)
            timing_data_list.append({"statistics": {}})
            batch_sizes.append(None)
            max_doc_lengths.append(None)
            continue
        timing_data_list.append(load_timing_data(tf))
        bs, mdl = get_metrics_info(tf, args.batch_size)
        batch_sizes.append(bs)
        max_doc_lengths.append(mdl)
        model_name = extract_model_name(tf)
        print(f"  {model_name}: max_doc_length={mdl}, batch_size={bs}")

    if args.per_item:
        print(f"Dividing statistics by batch size (per-item mode)")

    # Create workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for key in args.keys:
        tab_name = shorten_key_for_tab(key)
        ws = wb.create_sheet(title=tab_name)
        write_sheet(
            ws, key, args.timing_files, timing_data_list,
            batch_sizes if args.per_item else [None] * len(batch_sizes),
            max_doc_lengths,
            args.stats
        )

    wb.save(args.output)
    print(f"\nSaved to {args.output}")
    print(f"  {len(args.keys)} tab(s): {', '.join(shorten_key_for_tab(k) for k in args.keys)}")
    print(f"  {len(args.timing_files)} model(s) compared")


if __name__ == "__main__":
    main()